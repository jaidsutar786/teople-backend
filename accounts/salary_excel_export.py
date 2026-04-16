from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import MonthlySalary

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_monthly_salary_excel(request):
    """Download monthly salary data as Excel"""
    if request.user.role not in ['admin', 'manager']:
        return Response({"error": "Unauthorized"}, status=403)
    
    try:
        month = int(request.GET.get('month', datetime.now().month))
        year = int(request.GET.get('year', datetime.now().year))
        
        salaries = MonthlySalary.objects.filter(
            month=month,
            year=year
        ).select_related('employee').order_by('employee__first_name')
        
        if not salaries.exists():
            return Response({"error": "No salary records found"}, status=404)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{datetime(year, month, 1).strftime('%B %Y')}"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        headers = [
            'Sr No', 'Employee ID', 'Employee Name', 'Department', 'Position',
            'Present Days', 'Half Days', 'Leave Days', 'WFH Days', 'Comp Off Days',
            'Total Working Days', 'Gross Salary', 'Professional Tax', 'Final Salary',
            'Paid Leave Used', 'Unpaid Leave', 'Salary Cut Days', 'Generated Date'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, salary in enumerate(salaries, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=salary.employee.id)
            ws.cell(row=idx, column=3, value=f"{salary.employee.first_name} {salary.employee.last_name}")
            ws.cell(row=idx, column=4, value=salary.employee.department or 'N/A')
            ws.cell(row=idx, column=5, value=salary.employee.position or 'N/A')
            ws.cell(row=idx, column=6, value=salary.present_days)
            ws.cell(row=idx, column=7, value=salary.half_days)
            ws.cell(row=idx, column=8, value=salary.leave_days)
            ws.cell(row=idx, column=9, value=salary.wfh_days)
            ws.cell(row=idx, column=10, value=salary.comp_off_days)
            ws.cell(row=idx, column=11, value=salary.total_working_days)
            ws.cell(row=idx, column=12, value=float(salary.gross_monthly_salary))
            ws.cell(row=idx, column=13, value=float(salary.professional_tax))
            ws.cell(row=idx, column=14, value=float(salary.final_salary))
            ws.cell(row=idx, column=15, value=float(salary.paid_leave_used) if hasattr(salary, 'paid_leave_used') else 0)
            ws.cell(row=idx, column=16, value=float(salary.unpaid_leave_used) if hasattr(salary, 'unpaid_leave_used') else 0)
            ws.cell(row=idx, column=17, value=float(salary.salary_cut_days) if hasattr(salary, 'salary_cut_days') else 0)
            ws.cell(row=idx, column=18, value=salary.generated_at.strftime('%Y-%m-%d'))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Salary_Report_{datetime(year, month, 1).strftime('%B_%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return Response({"error": str(e)}, status=500)
